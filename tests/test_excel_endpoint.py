"""Tests for POST /generate/excel endpoint and excel_service."""

from io import BytesIO

import openpyxl

_XLSX_MAGIC = b"PK\x03\x04"  # ZIP/XLSX starts with PK


def _create_client(tmp_path, monkeypatch):
    monkeypatch.setenv("DECISIONDOC_PROVIDER", "mock")
    monkeypatch.setenv("DATA_DIR", str(tmp_path))
    monkeypatch.setenv("DECISIONDOC_ENV", "dev")
    monkeypatch.setenv("DECISIONDOC_MAINTENANCE", "0")
    monkeypatch.delenv("DECISIONDOC_API_KEY", raising=False)
    monkeypatch.delenv("DECISIONDOC_API_KEYS", raising=False)
    from fastapi.testclient import TestClient
    from app.main import create_app
    return TestClient(create_app())


def test_build_excel_returns_valid_bytes():
    from app.services.excel_service import build_excel
    docs = [{"doc_type": "adr", "markdown": "# 제목\n\n내용\n- 항목1\n- 항목2"}]
    result = build_excel(docs, title="테스트")
    assert result[:4] == _XLSX_MAGIC


def test_build_excel_multiple_docs():
    from app.services.excel_service import build_excel
    docs = [
        {"doc_type": "adr", "markdown": "# ADR\n내용"},
        {"doc_type": "onepager", "markdown": "# 원페이저\n## 요약\n내용"},
    ]
    result = build_excel(docs, title="멀티 문서")
    assert result[:4] == _XLSX_MAGIC


def test_excel_endpoint_returns_200(tmp_path, monkeypatch):
    client = _create_client(tmp_path, monkeypatch)
    res = client.post("/generate/excel", json={"title": "Excel 테스트", "goal": "검증"})
    assert res.status_code == 200


def test_excel_content_type(tmp_path, monkeypatch):
    client = _create_client(tmp_path, monkeypatch)
    res = client.post("/generate/excel", json={"title": "t", "goal": "g"})
    assert "spreadsheetml" in res.headers["content-type"]


def test_excel_content_disposition(tmp_path, monkeypatch):
    client = _create_client(tmp_path, monkeypatch)
    res = client.post("/generate/excel", json={"title": "한글제목", "goal": "g"})
    assert "attachment" in res.headers.get("content-disposition", "")


def test_excel_endpoint_returns_xlsx_magic(tmp_path, monkeypatch):
    client = _create_client(tmp_path, monkeypatch)
    res = client.post("/generate/excel", json={"title": "t", "goal": "g"})
    assert res.content[:4] == _XLSX_MAGIC


# ---------------------------------------------------------------------------
# Multi-sheet / metadata / boundary-case coverage
# ---------------------------------------------------------------------------


def test_build_excel_creates_cover_and_summary_and_doc_sheets():
    """Workbook should contain 표지 (cover) + 요약 (summary) + one sheet per doc_type."""
    from app.services.excel_service import build_excel

    docs = [
        {"doc_type": "adr", "markdown": "# ADR\n내용"},
        {"doc_type": "onepager", "markdown": "# 원페이저\n## 요약\n내용"},
    ]
    result = build_excel(docs, title="멀티 시트 검증")
    wb = openpyxl.load_workbook(BytesIO(result))

    assert wb.sheetnames[0] == "표지"
    assert wb.sheetnames[1] == "요약"
    # 4 sheets total: cover + summary + 2 doc sheets.
    assert len(wb.sheetnames) == 4
    # doc_type sheets use humanized Korean labels, not the raw doc_type.
    assert "adr" not in wb.sheetnames
    assert "onepager" not in wb.sheetnames


def test_build_excel_cover_sheet_contains_title_and_doc_count():
    from app.services.excel_service import build_excel

    docs = [{"doc_type": "adr", "markdown": "# ADR\n내용"}]
    result = build_excel(docs, title="표지 검증 제목")
    wb = openpyxl.load_workbook(BytesIO(result))
    cover_ws = wb["표지"]

    all_values = [
        str(cell.value) for row in cover_ws.iter_rows() for cell in row if cell.value is not None
    ]
    joined = " ".join(all_values)
    assert "표지 검증 제목" in joined
    assert "1개" in joined  # doc count


def test_build_excel_summary_sheet_has_metrics_header_and_rows():
    from app.services.excel_service import build_excel

    docs = [
        {"doc_type": "adr", "markdown": "# ADR\n| a | b |\n| --- | --- |\n| 1 | 2 |\n- 목록1\n- 목록2"},
        {"doc_type": "onepager", "markdown": "# 원페이저\n내용"},
    ]
    result = build_excel(docs, title="요약 시트 검증")
    wb = openpyxl.load_workbook(BytesIO(result))
    summary_ws = wb["요약"]

    header_row = [cell.value for cell in next(summary_ws.iter_rows(min_row=1, max_row=1))]
    assert "문서 유형" in header_row
    assert "표 수" in header_row
    assert "목록 수" in header_row

    # Row 2 corresponds to the first doc (adr): 1 table, 2 bullets.
    row2 = [cell.value for cell in next(summary_ws.iter_rows(min_row=2, max_row=2))]
    assert row2[3] == 1  # table count
    assert row2[4] == 2  # bullet count


def test_build_excel_empty_docs_list_still_produces_valid_workbook():
    from app.services.excel_service import build_excel

    result = build_excel([], title="빈 문서")
    assert result[:4] == _XLSX_MAGIC

    wb = openpyxl.load_workbook(BytesIO(result))
    assert "표지" in wb.sheetnames
    assert "요약" in wb.sheetnames
    # No doc sheets beyond cover + summary.
    assert len(wb.sheetnames) == 2


def test_build_excel_handles_markdown_special_characters():
    from app.services.excel_service import build_excel

    markdown = (
        "# 특수문자 테스트\n"
        "**볼드** 텍스트와 `코드` 그리고 & < > \" ' 문자\n"
        "- 목록 항목 with **bold** and 특수기호 %^&*()\n"
        "| 컬럼A | 컬럼B |\n"
        "| --- | --- |\n"
        "| **굵게** | 일반 |\n"
    )
    docs = [{"doc_type": "adr", "markdown": markdown}]
    result = build_excel(docs, title="특수문자 & <검증>")
    assert result[:4] == _XLSX_MAGIC

    wb = openpyxl.load_workbook(BytesIO(result))
    doc_sheet_name = [name for name in wb.sheetnames if name not in ("표지", "요약")][0]
    ws = wb[doc_sheet_name]
    all_values = [
        str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None
    ]
    joined = " ".join(all_values)
    # Bold markers should be stripped, not left as literal ** in cell text.
    assert "**" not in joined
    assert "볼드" in joined
    assert "굵게" in joined


def test_build_excel_truncates_cell_text_beyond_excel_limit():
    """A single markdown paragraph longer than Excel's 32,767-char cell limit
    must be clamped rather than raising or corrupting the workbook."""
    from app.services.excel_service import build_excel
    from app.services.excel_service import _EXCEL_CELL_CHAR_LIMIT

    huge_line = "가" * (_EXCEL_CELL_CHAR_LIMIT + 5000)
    docs = [{"doc_type": "adr", "markdown": f"# 제목\n{huge_line}"}]
    result = build_excel(docs, title="긴 텍스트 검증")
    assert result[:4] == _XLSX_MAGIC

    wb = openpyxl.load_workbook(BytesIO(result))
    doc_sheet_name = [name for name in wb.sheetnames if name not in ("표지", "요약")][0]
    ws = wb[doc_sheet_name]
    max_len = max(
        (len(str(cell.value)) for row in ws.iter_rows() for cell in row if cell.value is not None),
        default=0,
    )
    assert max_len <= _EXCEL_CELL_CHAR_LIMIT


def test_build_excel_sheet_names_are_unique_and_valid_for_duplicate_doc_types():
    """Duplicate/invalid doc_type values must not collide or contain forbidden chars."""
    from app.services.excel_service import build_excel

    docs = [
        {"doc_type": "adr", "markdown": "# ADR 1\n내용1"},
        {"doc_type": "adr", "markdown": "# ADR 2\n내용2"},
        {"doc_type": "weird/type:name*", "markdown": "# 이상한 타입\n내용"},
    ]
    result = build_excel(docs, title="시트 이름 검증")
    wb = openpyxl.load_workbook(BytesIO(result))

    doc_sheets = [name for name in wb.sheetnames if name not in ("표지", "요약")]
    assert len(doc_sheets) == 3
    assert len(set(doc_sheets)) == 3  # all unique
    for name in doc_sheets:
        assert len(name) <= 31
        assert not any(ch in name for ch in "\\/*?:[]")


def test_build_excel_long_doc_type_list_sheet_names_stay_within_limit():
    from app.services.excel_service import build_excel

    docs = [
        {"doc_type": f"very_long_document_type_name_number_{i}", "markdown": f"# 문서 {i}\n내용"}
        for i in range(3)
    ]
    result = build_excel(docs, title="긴 문서유형명 검증")
    wb = openpyxl.load_workbook(BytesIO(result))
    doc_sheets = [name for name in wb.sheetnames if name not in ("표지", "요약")]
    assert len(doc_sheets) == 3
    for name in doc_sheets:
        assert len(name) <= 31
